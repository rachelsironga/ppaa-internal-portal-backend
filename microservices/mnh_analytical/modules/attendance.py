import os
import logging
from datetime import datetime
from django.db import transaction
from django.db.models import Q, Sum
from django.utils import timezone
from rest_framework.exceptions import NotFound
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook

from microservices.mnh_analytical.models import Attendance, PatientAttendance, Clinic, PaymentMode
from microservices.mnh_analytical.serializers import (
    AttendanceSerializer, AttendanceListSerializer, AttendanceDetailSerializer
)
from mnh_approval.pagination import CustomPagination
from mnh_approval.response_codes import CustomResponse, STATUS_CODES
from utils.permissions import HasMethodPermission

logger = logging.getLogger(__name__)


class AttendanceView(APIView):
    permission_classes = [IsAuthenticated, HasMethodPermission]
    serializer_class = AttendanceSerializer
    required_permissions = {
        "get": ["view_attendance"],
        "post": ["add_attendance", "change_attendance"],
        "put": ["change_attendance"],
        "patch": ["change_attendance"],
        "delete": ["delete_attendance"]
    }

    def get(self, request, uid=None):
        try:
            if uid:
                attendance = Attendance.objects.using('analytical').filter(uid=uid, is_deleted=False).first()
                if not attendance:
                    raise NotFound("Attendance not found")
                return CustomResponse.success(data=AttendanceDetailSerializer(attendance).data)

            search_query = request.GET.get('search', '').strip()
            date_from = request.GET.get('date_from', '').strip()
            date_to = request.GET.get('date_to', '').strip()
            attendances = Attendance.objects.filter(is_deleted=False)

            if search_query:
                attendances = attendances.filter(
                    Q(notes__icontains=search_query)
                )

            if date_from:
                attendances = attendances.filter(date__gte=date_from)

            if date_to:
                attendances = attendances.filter(date__lte=date_to)

            if attendances.exists():
                return CustomPagination.paginate(view_class=self, results=attendances, request=request)

            return CustomResponse.errors(message="Attendances not found", data=[])
        except Exception as e:
            return CustomResponse.server_error(message=f'Failed to Retrieve Attendances: {str(e)}')

    def post(self, request):
        try:
            with transaction.atomic():
                serializer = self.serializer_class(data=request.data, context={'request': request})

                if serializer.is_valid():
                    serializer.save()
                    return CustomResponse.success(data=serializer.data)

                return CustomResponse.errors(
                    message="Validation Failed, Please Try Again",
                    data=serializer.errors,
                    code=STATUS_CODES["VALIDATION_ERROR"],
                )

        except Exception as e:
            return CustomResponse.server_error(message=f'Failed to Create Attendance: {str(e)}')

    def put(self, request, uid):
        try:
            with transaction.atomic():
                try:
                    instance = Attendance.objects.get(uid=uid, is_deleted=False)
                except Attendance.DoesNotExist:
                    return CustomResponse.errors(message="Attendance not found")

                serializer = self.serializer_class(instance, data=request.data, context={'request': request})

                if serializer.is_valid():
                    serializer.save()
                    return CustomResponse.success(data=serializer.data)

                return CustomResponse.errors(
                    message="Validation Failed, Please Try Again",
                    data=serializer.errors,
                    code=STATUS_CODES["VALIDATION_ERROR"],
                )

        except Exception as e:
            return CustomResponse.server_error(message=f'Failed to Update Attendance: {str(e)}')

    def patch(self, request, uid):
        try:
            with transaction.atomic():
                try:
                    instance = Attendance.objects.get(uid=uid, is_deleted=False)
                except Attendance.DoesNotExist:
                    return CustomResponse.errors(message="Attendance not found")

                serializer = self.serializer_class(
                    instance, data=request.data, partial=True, context={'request': request}
                )

                if serializer.is_valid():
                    serializer.save()
                    return CustomResponse.success(data=serializer.data)

                return CustomResponse.errors(
                    message="Validation Failed, Please Try Again",
                    data=serializer.errors,
                    code=STATUS_CODES["VALIDATION_ERROR"],
                )

        except Exception as e:
            return CustomResponse.server_error(message=f'Failed to Update Attendance: {str(e)}')

    def delete(self, request, uid):
        try:
            with transaction.atomic():
                attendance = Attendance.objects.filter(uid=uid, is_deleted=False).first()
                if not attendance:
                    return CustomResponse.errors(message="Attendance Not Found or Already Deleted")

                attendance.is_deleted = True
                attendance.deleted_at = datetime.now()
                attendance.deleted_by = request.user
                attendance.save()
                return CustomResponse.success(message='Attendance deleted successfully')

        except Exception as e:
            return CustomResponse.server_error(message="Something went wrong While Deleting Attendance")


class AttendanceSummaryView(APIView):
    permission_classes = [IsAuthenticated, HasMethodPermission]
    required_permissions = {
        "get": ["view_attendance"]
    }

    def get(self, request):
        try:
            date_from = request.GET.get('date_from', '').strip()
            date_to = request.GET.get('date_to', '').strip()

            attendances = Attendance.objects.filter(is_deleted=False)

            if date_from:
                attendances = attendances.filter(date__gte=date_from)

            if date_to:
                attendances = attendances.filter(date__lte=date_to)

            summary = attendances.aggregate(
                total_new_patients=Sum('total_new_patients'),
                total_follow_up_patients=Sum('total_follow_up_patients'),
                grand_total_patients=Sum('grand_total_patients')
            )

            data = {
                'total_attendances': attendances.count(),
                'total_new_patients': summary['total_new_patients'] or 0,
                'total_follow_up_patients': summary['total_follow_up_patients'] or 0,
                'grand_total_patients': summary['grand_total_patients'] or 0,
                'date_range_start': date_from or None,
                'date_range_end': date_to or None
            }

            return CustomResponse.success(data=data)
        except Exception as e:
            return CustomResponse.server_error(message=f'Failed to Retrieve Summary: {str(e)}')


def safe_int(value, default=0):
    """Safely convert a value to integer"""
    if value is None:
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return default
        if cleaned.isdigit():
            return int(cleaned)
        if cleaned.replace('.', '', 1).isdigit():
            try:
                return int(float(cleaned))
            except (ValueError, TypeError):
                return default
        if ',' in cleaned and cleaned.replace(',', '').isdigit():
            return int(cleaned.replace(',', ''))
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def get_or_create_clinic(clinic_name, user, report_date):
    """Get or create a clinic with proper validation"""
    clinic_name = clinic_name.strip()
    if not clinic_name or len(clinic_name) < 2:
        raise ValueError(f"Invalid clinic name: {clinic_name}")
    
    clinic_code = (
        clinic_name.upper()
        .replace(' ', '_')
        .replace('-', '_')
        .replace('/', '_')
        .replace('\\', '_')
        .replace('(', '')
        .replace(')', '')
        .replace("'", "")
        .replace('"', '')
    )
    while '__' in clinic_code:
        clinic_code = clinic_code.replace('__', '_')
    clinic_code = clinic_code.strip('_')
    
    existing = Clinic.objects.filter(
        Q(name__iexact=clinic_name) | Q(code__iexact=clinic_code)
    ).first()
    
    if existing:
        return existing
    
    clinic = Clinic.objects.create(
        name=clinic_name,
        code=clinic_code,
        created_by=user,
        description=f"Auto-created from attendance report on {report_date}"
    )
    logger.info(f"Created new clinic: {clinic_name} ({clinic_code})")
    return clinic


def get_or_create_payment_mode(payment_name, user):
    """Get or create a payment mode"""
    payment_name = payment_name.strip()
    if not payment_name or len(payment_name) < 2:
        raise ValueError(f"Invalid payment mode: {payment_name}")
    
    payment_code = (
        payment_name.upper()
        .replace(' ', '_')
        .replace('-', '_')
        .replace('/', '_')
    )
    while '__' in payment_code:
        payment_code = payment_code.replace('__', '_')
    payment_code = payment_code.strip('_')
    
    existing = PaymentMode.objects.filter(
        Q(name__iexact=payment_name) | Q(code__iexact=payment_code)
    ).first()
    
    if existing:
        return existing
    
    payment = PaymentMode.objects.create(
        name=payment_name,
        code=payment_code,
        created_by=user,
        description=f"Auto-created from attendance report"
    )
    logger.info(f"Created new payment mode: {payment_name} ({payment_code})")
    return payment


class AttendanceUploadView(APIView):
    permission_classes = [IsAuthenticated, HasMethodPermission]
    parser_classes = [MultiPartParser, FormParser]
    required_permissions = {
        "post": ["add_attendance", "upload_attendance"]
    }

    def post(self, request):
        try:
            date_str = request.data.get('date')
            report_file = request.FILES.get('attendance_report')
            process_now = request.data.get('process_now', 'false').lower() == 'true'
            
            if not date_str:
                return CustomResponse.errors(message="Date is required")
            if not report_file:
                return CustomResponse.errors(message="Attendance report file is required")
            
            try:
                attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return CustomResponse.errors(message="Invalid date format. Use YYYY-MM-DD")
            
            existing = Attendance.objects.filter(date=attendance_date, is_deleted=False).first()
            if existing:
                return CustomResponse.errors(
                    message=f"Attendance record for {attendance_date} already exists",
                    data={'existing_uid': str(existing.uid)}
                )
            
            with transaction.atomic():
                attendance = Attendance.objects.create(
                    date=attendance_date,
                    attendance_report=report_file,
                    created_by=request.user,
                    notes=request.data.get('notes', '')
                )
                
                if process_now:
                    result = self.process_report(attendance, request.user)
                    return CustomResponse.success(
                        data=AttendanceDetailSerializer(attendance).data,
                        message=f"Attendance uploaded and processed. {result['message']}"
                    )
                
                return CustomResponse.success(
                    data=AttendanceSerializer(attendance).data,
                    message="Attendance uploaded successfully. Processing pending."
                )
                
        except Exception as e:
            logger.exception("Error uploading attendance")
            return CustomResponse.server_error(message=f'Failed to Upload Attendance: {str(e)}')

    def process_report(self, attendance, user):
        """Process the uploaded Excel report"""
        total_columns = 0
        success_columns = 0
        failed_columns = 0
        total_new = 0
        total_follow_up = 0
        
        try:
            file_path = attendance.attendance_report.path
            if not os.path.exists(file_path):
                raise FileNotFoundError("Report file not found")
            
            wb = load_workbook(filename=file_path, data_only=True)
            sheet = wb.active
            
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) < 3:
                raise ValueError("Excel file must have at least 3 rows (headers + data)")
            
            payment_header = rows[0]
            sub_header = rows[1]
            
            payment_modes = []
            col = 1
            while col < len(payment_header):
                payment_name = payment_header[col]
                if payment_name and str(payment_name).strip():
                    payment_name = str(payment_name).strip()
                    try:
                        payment = get_or_create_payment_mode(payment_name, user)
                        payment_modes.append({
                            'payment': payment,
                            'fup_col': col,
                            'new_col': col + 1
                        })
                        total_columns += 1
                    except Exception as e:
                        logger.warning(f"Failed to create payment mode {payment_name}: {e}")
                        failed_columns += 1
                col += 2
            
            for row in rows[2:]:
                clinic_name = row[0] if row[0] else None
                if not clinic_name or not str(clinic_name).strip():
                    continue
                
                clinic_name = str(clinic_name).strip()
                try:
                    clinic = get_or_create_clinic(clinic_name, user, attendance.date)
                except Exception as e:
                    logger.warning(f"Failed to create clinic {clinic_name}: {e}")
                    continue
                
                for pm in payment_modes:
                    try:
                        fup_val = safe_int(row[pm['fup_col']] if pm['fup_col'] < len(row) else 0)
                        new_val = safe_int(row[pm['new_col']] if pm['new_col'] < len(row) else 0)
                        
                        if fup_val > 0 or new_val > 0:
                            PatientAttendance.objects.create(
                                attendance=attendance,
                                clinic=clinic,
                                payment=pm['payment'],
                                follow_up_patients=fup_val,
                                new_patients=new_val,
                                total_patients=fup_val + new_val,
                                created_by=user
                            )
                            total_new += new_val
                            total_follow_up += fup_val
                            success_columns += 1
                    except Exception as e:
                        logger.warning(f"Failed to create patient attendance: {e}")
                        failed_columns += 1
            
            attendance.total_new_patients = total_new
            attendance.total_follow_up_patients = total_follow_up
            attendance.grand_total_patients = total_new + total_follow_up
            attendance.processed_date = timezone.now().date()
            attendance.processed_by = user
            attendance.total_column = total_columns
            attendance.success_colums = success_columns
            attendance.failed_colums = failed_columns
            attendance.save()
            
            return {
                'success': True,
                'message': f"Processed {success_columns} records. New: {total_new}, Follow-up: {total_follow_up}, Total: {total_new + total_follow_up}"
            }
            
        except Exception as e:
            logger.exception("Error processing report")
            raise


class ProcessAttendanceView(APIView):
    permission_classes = [IsAuthenticated, HasMethodPermission]
    required_permissions = {
        "post": ["process_attendance", "change_attendance"]
    }

    def post(self, request, uid):
        try:
            attendance = Attendance.objects.filter(uid=uid, is_deleted=False).first()
            if not attendance:
                return CustomResponse.errors(message="Attendance not found")
            
            if attendance.processed_date:
                return CustomResponse.errors(
                    message=f"Attendance already processed on {attendance.processed_date}"
                )
            
            if not attendance.attendance_report:
                return CustomResponse.errors(message="No attendance report file to process")
            
            upload_view = AttendanceUploadView()
            result = upload_view.process_report(attendance, request.user)
            
            return CustomResponse.success(
                data=AttendanceDetailSerializer(attendance).data,
                message=result['message']
            )
            
        except Exception as e:
            logger.exception("Error processing attendance")
            return CustomResponse.server_error(message=f'Failed to Process Attendance: {str(e)}')
